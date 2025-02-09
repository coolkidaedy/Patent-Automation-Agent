import openpyxl
import json
from transformers import pipeline


model_name = "google/pegasus-xsum"  # PEGASUS model is optimized for concise summaries


summarizer = pipeline("summarization", model=model_name)


filler_words = {"is", "the", "and", "a", "an", "of", "to", "in", "on", "for", "at", "by", "with", "from", "as"}

def read_and_assign_summary_title(xlsx_file, output_json_file):
    
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb.active
    
    data = []
    
    
    total_rows = sheet.max_row - 1  # Subtract 1 for the header row

    
    for idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  
        description = row[0]
        
        
        if description:
            try:
               
                summary = summarizer(description, max_length=50, min_length=30, do_sample=False)
                summary_text = summary[0]['summary_text']
                
                
                cleaned_title_words = [word for word in summary_text.split() if word.lower() not in filler_words]
                
                
                concise_title = ' '.join(cleaned_title_words)
                
               
                if len(concise_title.split()) > 12:  # This is roughly 70 characters or fewer
                    concise_title = ' '.join(concise_title.split()[:12]) + "..."
                
            except Exception as e:
                concise_title = f"Error generating summary: {e}"
        else:
            concise_title = "No description provided"
        
        # Store the description and title in the list
        data.append({
            'description': description,
            'title': concise_title
        })
        
        
        if (idx - 1) % 10 == 0:  # Print progress every 10 rows
            progress = (idx - 1) / total_rows * 100
            print(f"Processing row {idx - 1}/{total_rows} ({progress:.2f}% complete)")
    
    
    with open(output_json_file, 'w') as json_file:
        json.dump(data, json_file, indent=4)

    print(f"Data has been saved to {output_json_file}")


read_and_assign_summary_title('spi.xlsx', 'output_titles_and_descriptions.json')

